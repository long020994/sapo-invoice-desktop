import csv
import json
import re
import sqlite3
from copy import copy
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def round_vnd(value):
    return int(Decimal(str(value or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def clean_sapo_reference(value):
    """Mã đơn nhập Sapo chỉ dùng chữ số để không vướng ký tự đặc biệt."""
    return "".join(re.findall(r"\d+", str(value or "")))[:100]


class InvoiceStorage:
    def __init__(self, path):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.initialize()

    def connect(self):
        connection = sqlite3.connect(self.path)
        connection.row_factory = sqlite3.Row
        return connection

    def initialize(self):
        with self.connect() as connection:
            connection.executescript("""
                CREATE TABLE IF NOT EXISTS invoices (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    created_at TEXT NOT NULL,
                    supplier_name TEXT,
                    invoice_number TEXT,
                    invoice_date TEXT,
                    filenames TEXT,
                    item_count INTEGER NOT NULL,
                    total_quantity REAL NOT NULL,
                    grand_total REAL NOT NULL,
                    status TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS invoice_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    invoice_id INTEGER NOT NULL REFERENCES invoices(id) ON DELETE CASCADE,
                    original_name TEXT,
                    sapo_name TEXT,
                    variant_id INTEGER,
                    sku TEXT,
                    barcode TEXT,
                    quantity REAL NOT NULL,
                    unit_price REAL NOT NULL,
                    line_total REAL NOT NULL,
                    confidence REAL,
                    matched INTEGER NOT NULL,
                    raw_json TEXT NOT NULL
                );
            """)

    def save_invoice(self, items, status="Đã duyệt"):
        if not items:
            raise ValueError("Không có sản phẩm để lưu")
        summary = items[0].get("invoice_summary") or {}
        filenames = sorted({str(item.get("invoice_filename") or "") for item in items})
        total_quantity = sum(float(item.get("qty") or 0) for item in items)
        grand_total = sum(float(item.get("price") or 0) * float(item.get("qty") or 0) for item in items)
        with self.connect() as connection:
            cursor = connection.execute(
                """INSERT INTO invoices
                (created_at, supplier_name, invoice_number, invoice_date, filenames,
                 item_count, total_quantity, grand_total, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    datetime.now().isoformat(timespec="seconds"),
                    summary.get("supplier_name", ""),
                    summary.get("invoice_number", ""),
                    summary.get("invoice_date", ""),
                    " | ".join(filenames),
                    len(items), total_quantity, grand_total, status,
                ),
            )
            invoice_id = cursor.lastrowid
            for item in items:
                connection.execute(
                    """INSERT INTO invoice_items
                    (invoice_id, original_name, sapo_name, variant_id, sku, barcode,
                     quantity, unit_price, line_total, confidence, matched, raw_json)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        invoice_id, item.get("original_name", ""), item.get("sapo_name", ""),
                        item.get("variant_id"), item.get("sku", ""), item.get("barcode", ""),
                        float(item.get("qty") or 0), float(item.get("price") or 0),
                        float(item.get("qty") or 0) * float(item.get("price") or 0),
                        float(item.get("confidence") or 0), int(bool(item.get("matched"))),
                        json.dumps(item, ensure_ascii=False),
                    ),
                )
        return invoice_id

    def recent(self, limit=200):
        with self.connect() as connection:
            return connection.execute(
                "SELECT * FROM invoices ORDER BY id DESC LIMIT ?", (limit,)
            ).fetchall()


def export_sapo_excel(items, destination, template_path):
    """Điền dữ liệu vào đúng mẫu nhập đơn hàng do Sapo cung cấp."""
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError("Không tìm thấy file mẫu nhập hàng Sapo đi kèm ứng dụng")
    workbook = load_workbook(template_path)
    sheet = workbook[workbook.sheetnames[0]]
    start_row = 9
    required_last_row = start_row + max(len(items), 1) - 1
    current_last_row = max(sheet.max_row, required_last_row)
    source_row = start_row
    for row in range(start_row, current_last_row + 1):
        if row > sheet.max_row:
            sheet.insert_rows(row)
        for column in range(1, 14):
            cell = sheet.cell(row, column)
            source = sheet.cell(source_row, column)
            if row != source_row:
                cell._style = copy(source._style)
                if source.has_style:
                    cell.number_format = source.number_format
                cell.alignment = copy(source.alignment)
                cell.border = copy(source.border)
                cell.fill = copy(source.fill)
                cell.font = copy(source.font)
                cell.protection = copy(source.protection)
            cell.value = None
        sheet.row_dimensions[row].height = sheet.row_dimensions[source_row].height

    summary = items[0].get("invoice_summary") or {} if items else {}
    reference = clean_sapo_reference(summary.get("invoice_number", ""))
    # Một số hóa đơn không có số. Dùng dấu thời gian toàn chữ số để Sapo vẫn nhận.
    if not reference:
        reference = datetime.now().strftime("%Y%m%d%H%M%S%f")
    sheet["B1"] = reference
    sheet["B2"] = "AI-HOA-DON"
    sheet["B3"] = "Tạo và kiểm tra bởi Sapo Invoice Desktop"
    sheet["B4"] = reference
    discount_value = round_vnd(summary.get("manual_discount_value"))
    discount_type = str(summary.get("manual_discount_type") or "VND").upper()
    # Mẫu Sapo dành G1 cho % và G2 cho số tiền VND; E/F là nhãn cố định.
    sheet["G1"] = None
    sheet["G2"] = None
    if discount_type == "%":
        sheet["G1"] = float(summary.get("manual_discount_value") or 0)
    elif discount_value > 0:
        sheet["G2"] = discount_value
    # Giá xuất ra đã gồm VAT, nên không cộng thuế thêm lần nữa trong Sapo.
    sheet["F3"] = None

    for offset, item in enumerate(items):
        row = start_row + offset
        sheet.cell(row, 1, str(item.get("sku") or ""))
        sheet.cell(row, 2, str(item.get("barcode") or ""))
        sheet.cell(row, 3, str(item.get("sapo_name") or item.get("original_name") or ""))
        sheet.cell(row, 4, "")
        sheet.cell(row, 5, float(item.get("qty") or 0))
        sheet.cell(row, 6, "")
        sheet.cell(row, 7, "")
        sheet.cell(row, 8, "")
        invoice_price = round_vnd(item.get("price"))
        system_cost = round_vnd(item.get("system_cost"))
        # Sapo chỉ nhận giá nguyên. Để trống khi giá vốn không đổi để tránh cập nhật thừa.
        sheet.cell(row, 9, None if invoice_price == system_cost else invoice_price)
        sheet.cell(row, 10, "")
        sheet.cell(row, 11, "")
        sheet.cell(row, 12, "")
        sheet.cell(row, 13, "")
    for table in sheet.tables.values():
        table.ref = f"A1:M{max(required_last_row, 8)}"
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def export_product_updates(items, destination):
    """Xuất danh sách SKU/giá bán cần cập nhật để người dùng kiểm tra trước."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SKU và giá bán"
    headers = (
        "ID phiên bản", "Tên sản phẩm", "Barcode", "SKU mới",
        "Giá vốn cũ", "Giá nhập mới", "Giá bán hiện tại", "Giá bán mới", "Cần cập nhật",
    )
    sheet.append(headers)
    for item in items:
        old_cost = round_vnd(item.get("system_cost"))
        new_cost = round_vnd(item.get("price"))
        old_sale = round_vnd(item.get("system_sale_price"))
        new_sale = round_vnd(item.get("new_sale_price") or old_sale)
        sku_new = bool(item.get("generated_sku") or item.get("sku_changed"))
        sale_changed = new_sale != old_sale
        barcode_changed = bool(item.get("barcode_changed"))
        if not (sku_new or barcode_changed or sale_changed or new_cost != old_cost):
            continue
        reasons = []
        if sku_new:
            reasons.append("Thêm SKU")
        if barcode_changed:
            reasons.append("Đổi barcode")
        if new_cost != old_cost:
            reasons.append("Đổi giá vốn")
        if sale_changed:
            reasons.append("Đổi giá bán")
        sheet.append((
            item.get("variant_id"), item.get("sapo_name") or item.get("original_name"),
            item.get("barcode", ""), item.get("sku", ""), old_cost, new_cost,
            old_sale, new_sale, ", ".join(reasons),
        ))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="123B63")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = (16, 42, 18, 24, 15, 15, 18, 16, 28)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2, min_col=5, max_col=8):
        for cell in row:
            cell.number_format = '#,##0 "đ"'
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def export_sapo_product_csv(items, source_csv, destination):
    """Tạo file cập nhật sản phẩm Sapo từ chính file xuất danh mục mới nhất."""
    source_csv = Path(source_csv)
    if not source_csv.exists():
        raise FileNotFoundError("Không tìm thấy file danh sách sản phẩm Sapo")

    with source_csv.open("r", encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        headers = list(reader.fieldnames or [])
        required = {"Id phiên bản", "Mã SKU", "Barcode", "Giá", "Tên sản phẩm*"}
        missing = required.difference(headers)
        if missing:
            raise ValueError(
                "File sản phẩm Sapo thiếu cột bắt buộc: " + ", ".join(sorted(missing))
            )
        rows = list(reader)

    updates = {}
    for item in items:
        if not item.get("matched") or item.get("variant_id") is None:
            continue
        old_sale = round_vnd(item.get("system_sale_price"))
        new_sale = round_vnd(item.get("new_sale_price") or old_sale)
        sku_changed = bool(item.get("generated_sku") or item.get("sku_changed"))
        barcode_changed = bool(item.get("barcode_changed"))
        sale_changed = new_sale != old_sale
        if sku_changed or barcode_changed or sale_changed:
            updates[str(int(item["variant_id"]))] = {
                "sku": str(item.get("sku") or "").strip(),
                "barcode": str(item.get("barcode") or "").strip(),
                "sale": new_sale,
                "sku_changed": sku_changed,
                "barcode_changed": barcode_changed,
                "sale_changed": sale_changed,
            }
    if not updates:
        return None

    output_rows = []
    found = set()
    for row in rows:
        raw_variant_id = str(row.get("Id phiên bản") or "").strip()
        try:
            variant_id = str(int(float(raw_variant_id)))
        except ValueError:
            continue
        update = updates.get(variant_id)
        if not update:
            continue
        found.add(variant_id)
        edited = dict(row)
        if update["sku_changed"]:
            edited["Mã SKU"] = update["sku"]
        if update["barcode_changed"]:
            edited["Barcode"] = update["barcode"]
        if update["sale_changed"]:
            edited["Giá"] = str(update["sale"])
        output_rows.append(edited)

    missing_ids = sorted(set(updates).difference(found))
    if missing_ids:
        raise ValueError(
            f"File sản phẩm không chứa {len(missing_ids)} phiên bản cần cập nhật. "
            "Hãy xuất lại danh sách sản phẩm mới nhất từ Sapo."
        )

    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(output_rows)
    return destination


def export_bulk_skus_from_barcodes(source_csv, destination, database_destination=None):
    """Chỉ điền SKU còn trống bằng barcode; thêm hậu tố số khi mã bị trùng."""
    source_csv = Path(source_csv)
    with source_csv.open("r", encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        headers = list(reader.fieldnames or [])
        required = {"Tên sản phẩm*", "Mã SKU", "Barcode", "Giá", "Giá vốn", "Id phiên bản"}
        missing = required.difference(headers)
        if missing:
            raise ValueError(
                "File sản phẩm Sapo thiếu cột bắt buộc: " + ", ".join(sorted(missing))
            )
        rows = list(reader)

    # Mọi SKU đã có đều là dữ liệu cần bảo toàn.
    identifier_owners = {}
    for row_index, row in enumerate(rows):
        for field in ("Mã SKU", "Barcode"):
            identifier = str(row.get(field) or "").strip().casefold()
            if identifier:
                identifier_owners.setdefault(identifier, set()).add(row_index)

    def conflicts(identifier, row_index):
        return bool(identifier_owners.get(identifier.casefold(), set()).difference({row_index}))
    changed_rows = []
    duplicate_suffixes = 0
    skipped_no_barcode = 0
    preserved_existing_sku = 0
    for row_index, row in enumerate(rows):
        previous_sku = str(row.get("Mã SKU") or "").strip()
        if previous_sku:
            preserved_existing_sku += 1
            continue
        barcode = str(row.get("Barcode") or "").strip()
        if not barcode:
            skipped_no_barcode += 1
            continue
        candidate = barcode
        suffix = 1
        while conflicts(candidate, row_index):
            candidate = f"{barcode}{suffix}"
            suffix += 1
        if candidate != barcode:
            duplicate_suffixes += 1
        row["Mã SKU"] = candidate
        identifier_owners.setdefault(candidate.casefold(), set()).add(row_index)
        changed_rows.append(dict(row))

    if not changed_rows:
        raise ValueError("Không còn sản phẩm trống SKU và có barcode để cập nhật")

    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(changed_rows)

    if database_destination:
        def clean(value):
            return str(value or "").strip()

        def parse_price(value):
            try:
                return float(clean(value).replace(" ", "").replace(",", ""))
            except ValueError:
                return 0.0

        database = []
        for row in rows:
            try:
                variant_id = int(float(clean(row.get("Id phiên bản"))))
            except ValueError:
                continue
            name_parts = [clean(row.get("Tên sản phẩm*"))]
            for index in range(1, 4):
                value = clean(row.get(f"Giá trị thuộc tính {index}"))
                if value and value.casefold() != "default title":
                    name_parts.append(value)
            name = " - ".join(part for part in name_parts if part)
            if not name or not variant_id:
                continue
            price = parse_price(row.get("Giá"))
            cost = parse_price(row.get("Giá vốn"))
            database.append({
                "name": name,
                "sku": clean(row.get("Mã SKU")),
                "barcode": clean(row.get("Barcode")),
                "variant_id": variant_id,
                "price": price,
                "cost": cost,
                "prices": sorted({value for value in (price, cost) if value > 0}),
            })
        database_destination = Path(database_destination)
        database_destination.parent.mkdir(parents=True, exist_ok=True)
        temporary = database_destination.with_suffix(database_destination.suffix + ".tmp")
        temporary.write_text(json.dumps(database, ensure_ascii=False, indent=2), encoding="utf-8")
        temporary.replace(database_destination)

    return {
        "total_products": len(rows),
        "updated": len(changed_rows),
        "duplicate_suffixes": duplicate_suffixes,
        "skipped_no_barcode": skipped_no_barcode,
        "preserved_existing_sku": preserved_existing_sku,
        "output": str(destination),
    }
